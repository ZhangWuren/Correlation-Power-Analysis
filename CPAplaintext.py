from openpyxl import load_workbook
import xlsxwriter as xs


def load_data(filename):
    wb = load_workbook(filename)
    sheets = wb.worksheets
    sheet1 = sheets[0]
    return sheet1


if __name__ == "__main__":
    plaintext = load_data('plaintext.xlsx')
    workbook = xs.Workbook('plaintext2.xlsx')
    worksheet = workbook.add_worksheet('sheet1')
    for i in range(300):
        cell = plaintext.cell(i + 1, 1).value
        temp = cell.split()
        for j in range(16):
            worksheet.write(i, j, temp[j])

    workbook.close()
